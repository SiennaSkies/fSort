import sys
import re
import openpyxl

def process_file(filename):
    workbook = openpyxl.Workbook()
    
    sheet1 = workbook.active
    sheet1.title = "端口信息"
    sheet2 = workbook.create_sheet(title="网站信息")
    sheet3 = workbook.create_sheet(title="漏洞信息")
    sheet4 = workbook.create_sheet(title="主机信息")
    sheet5 = workbook.create_sheet(title="系统信息")
    sheet6 = workbook.create_sheet(title="POC信息")
    
    with open(filename, 'r', encoding='gbk', errors='ignore') as file:
        lines = file.readlines()
        for i, line in enumerate(lines):
            line = line.strip() 
            
            match_ip_port = re.match(r'(\d+\.\d+\.\d+\.\d+):(\d+)', line)
            if match_ip_port:
                ip = match_ip_port.group(1)
                port = match_ip_port.group(2)
                sheet1.append([ip, port])
            
            elif line.startswith("[*]"):
                if line.startswith("[*] WebTitle"):
                    url = line.replace("[*] WebTitle ", "").replace(" ","").split("code:")[0]
                    respond = line.replace("[*] WebTitle ", "").split("code:")[1].split("len:")[0]
                    title = line.replace("[*] WebTitle ", "").split("title:")[1]
                    sheet2.append([url, respond, title])

                elif line.startswith("[*] NetInfo"):
                    if i + 1 < len(lines) and i + 2 < len(lines):
                        IPaddress = lines[i + 1].strip().replace("[*]", "")
                        hostname = lines[i + 2].strip().replace("[->]", "")
                        sheet4.append([IPaddress, hostname])
                
                elif line.startswith("[*] NetBios"):
                    sheet5.append([line.replace("[*] NetBios ", "")])

            elif line.startswith("[+]"):
                if line.startswith("[+] InfoScan"):
                    sheet2.append([line])
                elif line.startswith("[+] PocScan"):
                    sheet6.append([line.replace("PocScan","")])
                else:
                    sheet3.append([line])
    
    output_filename = filename.replace('.txt', '.xlsx')
    workbook.save(output_filename)
    print(f"Excel文件已保存为 {output_filename}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("使用用法: python fSort.py <目标txt文件>")
        sys.exit(1)
    
    filename = sys.argv[1]
    process_file(filename)
